#!/usr/bin/env python3
"""
build_schedule_data.py — build faculty-schedule/data/<TERM>/sections.json from a
scraped schedule spreadsheet.

One source: the scrape. No enrollment manifest, no cross-repo dependency.

    python3 tools/build_schedule_data.py --inspect scrape.xlsx
    python3 tools/build_schedule_data.py scrape.xlsx \
        --term 26FA --term-label "Fall 2026" \
        --out faculty-schedule/data/26FA/sections.json

Accepts .xlsx, .xlsm, .csv, or .tsv.

Column names are matched loosely against the alias table below, so
"Section Start Time", "Start Time", "Begin Time", and "start" all resolve to the
same field. Run --inspect first: it prints which spreadsheet column mapped to
which field, which fields went unmatched, and a few fully parsed sections, and
writes nothing. If a column is unmatched, add its header to ALIASES and rerun —
that is the intended way to adapt this to a new export.

Rows are grouped by section code, so a section that lectures MW and labs F
should appear as TWO rows sharing a section code. Every row of a section
contributes one meeting pattern.
"""

import argparse, csv, json, os, re, sys
from collections import OrderedDict, defaultdict
from datetime import datetime, timezone

# ---------------------------------------------------------------- field aliases

# field -> candidate header names, lowercased with punctuation stripped.
# Put specific names before generic ones; the first match wins.
ALIASES = OrderedDict([
    ("section",     ["section", "sectioncode", "sectionid", "coursesection", "crsesection"]),
    ("crn",         ["crn", "coursereferencenumber", "refnumber"]),
    ("course",      ["course", "coursenumber", "coursecode", "subjectcourse", "catalog"]),
    ("title",       ["coursetitle", "sectiontitle", "title", "coursename", "longtitle"]),
    ("credits",     ["coursecredit", "credits", "credithours", "creditvalue", "cr", "units"]),
    ("instructor",  ["instructor", "instructors", "primaryinstructor", "faculty", "teacher",
                     "instructorname", "assignedinstructor"]),
    ("days",        ["meetingdays", "days", "daysofweek", "meetingpattern", "daysmet", "day"]),
    ("start",       ["sectionstarttime", "starttime", "begintime", "meetingstart", "start", "timestart"]),
    ("end",         ["sectionendtime", "endtime", "stoptime", "meetingend", "end", "timeend"]),
    ("building",    ["building", "bldg", "buildingcode", "buildingname"]),
    ("room",        ["room", "roomnumber", "roomno", "meetingroom", "buildingroom"]),
    ("campus",      ["campus", "location", "site", "campuscode"]),
    ("pot",         ["partofterm", "pot", "session", "sessioncode", "sessionlength"]),
    ("mode",        ["instructionalmethod", "deliverymethod", "mode", "method", "instructiontype",
                     "deliverymode", "format"]),
    ("mtype",       ["meetingtype", "scheduletype", "type", "component"]),
    ("startdate",   ["startdate", "sectionstartdate", "begindate", "partoftermstart"]),
    ("enddate",     ["enddate", "sectionenddate", "partoftermend"]),
    ("term",        ["term", "termcode", "academicterm"]),
    ("status",      ["status", "sectionstatus", "cancelled", "canceled"]),
])

DAY_TOKENS = {"M": "M", "MO": "M", "MON": "M", "MONDAY": "M",
              "T": "T", "TU": "T", "TUE": "T", "TUES": "T", "TUESDAY": "T",
              "W": "W", "WE": "W", "WED": "W", "WEDNESDAY": "W",
              "TH": "R", "R": "R", "THU": "R", "THUR": "R", "THURS": "R", "THURSDAY": "R",
              "F": "F", "FR": "F", "FRI": "F", "FRIDAY": "F",
              "SA": "S", "S": "S", "SAT": "S", "SATURDAY": "S",
              "SU": "U", "U": "U", "SUN": "U", "SUNDAY": "U"}

ONLINE_CAMPUS = {"DE", "ONL", "ONLINE", "WEB"}
HYBRID_CAMPUS = {"ONLH", "HYB"}
# Campus codes that describe a delivery mode rather than a place. The official
# faculty schedule form writes "Remote Live" where a room would go for ORLV.
CAMPUS_ROOM_LABEL = {"ORLV": "Remote Live"}
SYNC_ONLINE_CAMPUS = {"ORLV"}
VALID_MODES = {"inperson", "online-async", "online-sync", "hybrid"}
MODE_HINTS = [
    (r"asynchron|self.?paced|online.?async", "online-async"),
    (r"synchron|remote.?live|online.?live|zoom|virtual.?class", "online-sync"),
    (r"hybrid|blended|mixed", "hybrid"),
    (r"in.?person|face.?to.?face|f2f|traditional|on.?campus|lecture|lab", "inperson"),
]
NULLISH = {"", "TBA", "TBD", "NONE", "N/A", "NA", "-", "--", "STAFF", "ONLINE"}


def norm_header(h):
    return re.sub(r"[^a-z0-9]", "", str(h or "").lower())


def blank(v):
    return v is None or str(v).strip().upper() in NULLISH


def to_min(t):
    """'2:00 PM' / '14:00' / '1400' / datetime.time -> minutes past midnight."""
    if t is None or t == "":
        return None
    if hasattr(t, "hour") and hasattr(t, "minute"):
        return t.hour * 60 + t.minute
    s = str(t).strip().upper().replace(".", "")
    if s in NULLISH:
        return None
    m = re.match(r"^(\d{1,2}):(\d{2})(?::\d{2})?\s*(AM|PM)?$", s)
    if m:
        h, mi, ap = int(m.group(1)), int(m.group(2)), m.group(3)
    else:
        m2 = re.match(r"^(\d{3,4})\s*(AM|PM)?$", s)
        if not m2:
            return None
        digits = m2.group(1).zfill(4)
        h, mi, ap = int(digits[:2]), int(digits[2:]), m2.group(2)
    if ap == "PM" and h != 12:
        h += 12
    if ap == "AM" and h == 12:
        h = 0
    return h * 60 + mi if h < 24 and mi < 60 else None


def hhmm(mins):
    return None if mins is None else "%02d:%02d" % (mins // 60, mins % 60)


def parse_days(s):
    """'M, W' / 'MW' / 'MWF' / 'TuTh' / 'M-W-F' -> ['M','W',...]"""
    if blank(s):
        return []
    s = str(s).strip().upper()
    toks = re.split(r"[,\s/;|+-]+", s) if re.search(r"[,\s/;|+-]", s) else None
    if toks is None:
        toks, i = [], 0
        while i < len(s):                       # two-char first, so TH beats T
            if s[i:i + 2] in DAY_TOKENS:
                toks.append(s[i:i + 2]); i += 2
            elif s[i] in DAY_TOKENS:
                toks.append(s[i]); i += 1
            else:
                i += 1
    out = []
    for t in toks:
        k = DAY_TOKENS.get(t.strip())
        if k and k not in out:
            out.append(k)
    return out


def parse_instructors(v, name_style="auto"):
    """Split an instructor cell into names.

    This column mixes conventions in the same file:
        'Michael Qaissaunee'              one person, First Last
        'Qaissaunee, M'                   one person, Last, Initial
        'Hansen, P; Healy, J; Cole, C'    three people, semicolon separated
        'Cheryl Fencik, Gregory Augustino'  two people, comma separated
    The comma is the ambiguous one. A comma between two multi-token names
    separates people; a comma between two single tokens is Last, First.
    """
    if blank(v):
        return []
    raw = " ".join(str(v).split())

    if ";" in raw:
        parts = raw.split(";")
    elif re.search(r"\s(?:&|and)\s|/", raw):
        parts = re.split(r"\s(?:&|and)\s|/", raw)
    elif "," in raw:
        bits = [b.strip() for b in raw.split(",")]
        if all(len(b.split()) >= 2 for b in bits if b):
            parts = bits                       # 'Cheryl Fencik, Gregory Augustino'
        elif len(bits) > 2:
            parts = [", ".join(bits[i:i + 2]) for i in range(0, len(bits) - 1, 2)]
        else:
            parts = [raw]                      # 'Qaissaunee, M'
    else:
        parts = [raw]

    names = []
    for p in parts:
        p = " ".join(p.split()).strip(",;")
        if not p or p.upper() in NULLISH:
            continue
        if name_style in ("auto", "first-last") and re.match(r"^[^,]+,\s*[^,]+$", p):
            last, first = [x.strip() for x in p.split(",", 1)]
            p = first + " " + last
        names.append(p)
    return names


def canonicalize_names(sections, warn):
    """Merge 'M Qaissaunee' into 'Michael Qaissaunee' when it is unambiguous.

    Same person, two spellings, is the most damaging error this feed can carry:
    the picker lists them twice and each entry holds part of the schedule.
    """
    forms = {}
    for s in sections:
        for n in s["i"]:
            toks = n.split()
            if len(toks) < 2:
                continue
            surname = " ".join(toks[1:]).lower()
            forms.setdefault(surname, set()).add(n)

    rewrite = {}
    for surname, variants in forms.items():
        full = [v for v in variants if len(v.split()[0].rstrip(".")) > 1]
        short = [v for v in variants if len(v.split()[0].rstrip(".")) == 1]
        for sh in short:
            initial = sh.split()[0].rstrip(".").lower()
            hits = [f for f in full if f.split()[0][:1].lower() == initial]
            if len(hits) == 1:
                rewrite[sh] = hits[0]
            elif len(hits) > 1:
                warn(f"'{sh}' could be any of {hits} — left as its own entry")
    if rewrite:
        for s in sections:
            s["i"] = list(dict.fromkeys(rewrite.get(n, n) for n in s["i"]))
        for sh, full in sorted(rewrite.items())[:8]:
            warn(f"merged '{sh}' into '{full}'")
        if len(rewrite) > 8:
            warn(f"…and {len(rewrite) - 8} more name merges")
    return len(rewrite)


def parse_date(v):
    if blank(v):
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d-%b-%Y", "%b %d, %Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def parse_credits(v):
    if blank(v):
        return 0
    try:
        f = float(str(v).strip())
        return int(f) if f == int(f) else f
    except ValueError:
        return 0


def infer_mode(mode_raw, mtype_raw, campus, days, start):
    for raw in (mode_raw, mtype_raw):
        if not blank(raw):
            s = str(raw).strip()
            flat = s.lower().replace(" ", "").replace("_", "-")
            if flat in VALID_MODES:
                return flat
            for pat, mode in MODE_HINTS:
                if re.search(pat, s, re.I):
                    return mode
    c = (campus or "").upper()
    if c in SYNC_ONLINE_CAMPUS:
        return "online-sync"
    if c in ONLINE_CAMPUS:
        return "online-sync" if days else "online-async"
    if c in HYBRID_CAMPUS:
        return "hybrid"
    if not days and start is None:
        return "online-async"
    return "inperson"


def dept_of(course, section):
    m = re.match(r"^([A-Z]+)", (course or section or ""))
    return m.group(1) if m else "OTHER"


# ---------------------------------------------------------------- reading

def read_rows(path):
    """-> (headers, list of dicts keyed by original header)"""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".csv", ".tsv", ".txt"):
        delim = "\t" if ext == ".tsv" else ","
        with open(path, newline="", encoding="utf-8-sig") as f:
            rdr = csv.DictReader(f, delimiter=delim)
            return list(rdr.fieldnames or []), [dict(r) for r in rdr]
    try:
        import openpyxl
    except ImportError:
        sys.exit("Reading .xlsx needs openpyxl:  pip3 install openpyxl")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # This report ships two sheets: a consolidated course-level summary first,
    # then the section-level detail. Always take the section-level one.
    name = next((n for n in wb.sheetnames if re.search(r"individual\s*sect|section", n, re.I)),
                wb.sheetnames[-1] if len(wb.sheetnames) > 1 else wb.sheetnames[0])
    ws = wb[name]
    it = ws.iter_rows(values_only=True)
    headers = []
    for row in it:                                # skip leading blank or title rows
        vals = [c for c in row if c not in (None, "")]
        if len(vals) >= 3:
            headers = [str(c).strip() if c is not None else "" for c in row]
            break
    rows = []
    for row in it:
        if all(c in (None, "") for c in row):
            continue
        rows.append(dict(zip(headers, row)))
    return headers, rows


def map_columns(headers):
    """-> (field -> header, unmatched headers)"""
    norm = {}
    for h in headers:
        if h and norm_header(h) not in norm:
            norm[norm_header(h)] = h
    mapping, used = {}, set()
    for field, cands in ALIASES.items():
        for c in cands:
            if c in norm and norm[c] not in used:
                mapping[field] = norm[c]
                used.add(norm[c])
                break
    unmatched = [h for h in headers if h and h not in used]
    return mapping, unmatched


def get(row, mapping, field):
    h = mapping.get(field)
    return row.get(h) if h else None


# ---------------------------------------------------------------- build

def split_parts(v):
    """Comma-separated parallel list -> parts, empties preserved for alignment."""
    if v is None:
        return []
    return [p.strip() for p in str(v).split(",")]


def row_meetings(row, mapping):
    """One spreadsheet row -> zero or more meeting patterns.

    Handles both shapes:
      - one row per pattern      days="MW"        start="11:00 AM"
      - parallel packed lists    days="M W, F"    start=" 11:00 AM,  1:30 PM"
    Within a pattern days are space-separated; commas separate patterns.
    """
    dparts = split_parts(get(row, mapping, "days"))
    sparts = split_parts(get(row, mapping, "start"))
    eparts = split_parts(get(row, mapping, "end"))
    bparts = split_parts(get(row, mapping, "building"))
    rparts = split_parts(get(row, mapping, "room"))

    # "M, W" with a single start time is one pattern on two days, not two
    # patterns — collapse it rather than dropping the second day.
    if len(dparts) > 1 and len([x for x in sparts if x]) == 1:
        dparts = [" ".join(dparts)]

    n = max(len(dparts), len(sparts), 1)
    out = []
    for i in range(n):
        days = parse_days(dparts[i]) if i < len(dparts) else []
        st = to_min(sparts[i]) if i < len(sparts) else None
        en = to_min(eparts[i]) if i < len(eparts) else None
        if not days and st is None:
            continue                                   # async or padding
        bld = bparts[i] if i < len(bparts) else None
        rm = rparts[i] if i < len(rparts) else None
        label = " ".join(str(x).strip() for x in (bld, rm) if not blank(x)) or None
        if label is None:
            label = CAMPUS_ROOM_LABEL.get(str(get(row, mapping, "campus") or "").strip().upper())
        gap = label is None and (any(not blank(x) for x in rparts) or any(not blank(x) for x in bparts))
        out.append({"d": days, "s": hhmm(st), "e": hhmm(en), "rm": label,
                    "est": st is not None and en is None, "_room_gap": gap})
    return out



def build(rows, mapping, args, warn):
    groups = defaultdict(list)
    skipped = 0
    for r in rows:
        code = get(r, mapping, "section")
        if blank(code):
            skipped += 1
            continue
        groups[str(code).strip()].append(r)
    if skipped:
        warn(f"{skipped} rows had no section code and were skipped")

    sections = []
    for code, rs in sorted(groups.items()):
        first = rs[0]
        status = get(first, mapping, "status")
        if not blank(status) and re.search(r"cancel", str(status), re.I) and not args.keep_cancelled:
            continue

        course = get(first, mapping, "course")
        course = str(course).strip() if not blank(course) else re.sub(r"-[^-]+$", "", code)
        title = get(first, mapping, "title")
        title = str(title).strip() if not blank(title) else course
        campus = get(first, mapping, "campus")
        campus = str(campus).strip() if not blank(campus) else None
        pot = get(first, mapping, "pot")
        pot = str(pot).strip() if not blank(pot) else (args.default_pot or "15W")
        if args.pot and pot != args.pot:
            continue

        instructors, seen = [], set()
        for r in rs:
            for n in parse_instructors(get(r, mapping, "instructor"), args.names):
                if n not in seen:
                    seen.add(n)
                    instructors.append(n)

        meetings, no_end, part_room = [], False, False
        for r in rs:
            for m in row_meetings(r, mapping):
                if m["est"]:
                    no_end = True
                if m.pop("_room_gap", False):
                    part_room = True
                meetings.append(m)
        if no_end:
            warn(f"{code}: has a start time with no end time")
        if part_room:
            warn(f"{code}: fewer rooms listed than meeting patterns; the extras print blank")
        if len(rs) > 1 and not meetings:
            warn(f"{code}: {len(rs)} rows but no usable meeting times")

        mode = infer_mode(get(first, mapping, "mode"), get(first, mapping, "mtype"), campus,
                          meetings[0]["d"] if meetings else [],
                          to_min(meetings[0]["s"]) if meetings else None)

        crn = get(first, mapping, "crn")
        sections.append({
            "id": code,
            "c": course,
            "t": title,
            "cr": parse_credits(get(first, mapping, "credits")),
            "dept": dept_of(course, code),
            "campus": campus,
            "pot": pot,
            "mode": mode,
            "i": instructors,
            "m": meetings,
            "sd": parse_date(get(first, mapping, "startdate")),
            "ed": parse_date(get(first, mapping, "enddate")),
            "crn": None if blank(crn) else str(crn).strip(),
        })
    return sections


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("spreadsheet", nargs="+",
                    help="one or more schedule files (.xlsx, .xlsm, .csv, .tsv)")
    ap.add_argument("--pot-labels", default=None,
                    help="comma-separated part-of-term label per input file, e.g. 15W,11W,7A,7B")
    ap.add_argument("--out", default="sections.json")
    ap.add_argument("--term", default=None, help="e.g. 26FA")
    ap.add_argument("--term-label", default=None, help='e.g. "Fall 2026"')
    ap.add_argument("--pot", default=None, help="keep only this part of term, e.g. 15W")
    ap.add_argument("--default-pot", default="15W", help="used when the sheet has no part-of-term column")
    ap.add_argument("--names", default="auto", choices=["auto", "first-last", "as-is"],
                    help="auto/first-last flip 'Doe, Jane' to 'Jane Doe'; as-is leaves names alone")
    ap.add_argument("--calendar", default=None,
                    help="JSON with term start/end, closures, and swap days (see tools/calendar-26FA.json)")
    ap.add_argument("--keep-cancelled", action="store_true")
    ap.add_argument("--inspect", action="store_true", help="report column mapping and samples, write nothing")
    args = ap.parse_args()

    labels = [x.strip() for x in args.pot_labels.split(",")] if args.pot_labels else []
    if labels and len(labels) != len(args.spreadsheet):
        sys.exit(f"--pot-labels has {len(labels)} entries but {len(args.spreadsheet)} files were given.")

    warnings = []
    warn = warnings.append
    all_sections, seen, headers, mapping, unmatched = [], set(), [], {}, []

    for i, path in enumerate(args.spreadsheet):
        headers, rows = read_rows(path)
        mapping, unmatched = map_columns(headers)
        if "section" not in mapping:
            print(f"{path} columns: " + ", ".join(str(h) for h in headers if h))
            sys.exit("\nNo section-code column matched. Add its header to ALIASES['section'] and rerun.")
        file_args = argparse.Namespace(**vars(args))
        if labels:
            file_args.default_pot = labels[i]
        for sec in build(rows, mapping, file_args, warn):
            if sec["id"] in seen:
                warn(f"{sec['id']} appears in more than one file; keeping the first")
                continue
            seen.add(sec["id"])
            all_sections.append(sec)
        if len(args.spreadsheet) > 1:
            print(f"  {os.path.basename(path)}: {len(rows)} rows"
                  + (f" tagged {labels[i]}" if labels else ""))

    if args.inspect:
        print(f"{args.spreadsheet[0]}: {len(rows)} rows, {len(headers)} columns\n")
        print("MAPPED")
        for f in ALIASES:
            if f in mapping:
                print(f"  {f:<11} <- {mapping[f]}")
        missing = [f for f in ALIASES if f not in mapping]
        if missing:
            print("\nNOT FOUND (fine if the sheet genuinely lacks them)\n  " + ", ".join(missing))
        if unmatched:
            print("\nSPREADSHEET COLUMNS NOT USED\n  " + ", ".join(str(h) for h in unmatched))
        secs = all_sections
        canonicalize_names(secs, warn)
        multi = [s for s in secs if len(s["m"]) > 1]
        print(f"\n{len(secs)} sections · {len(multi)} with more than one meeting pattern")
        print("\nSAMPLE")
        shown, seen_ids = 0, set()
        for s in multi[:1] + secs:
            if s["id"] in seen_ids:
                continue
            seen_ids.add(s["id"])
            print("  " + json.dumps(s, ensure_ascii=False))
            shown += 1
            if shown == 3:
                break
        if warnings:
            print(f"\n{len(warnings)} warnings, first few:")
            for w in warnings[:5]:
                print("  ! " + w)
        return

    sections = all_sections
    merged = canonicalize_names(sections, warn)
    if not sections:
        sys.exit("No sections built. Run with --inspect to see how the columns mapped.")

    faculty = sorted({n for s in sections for n in s["i"]},
                     key=lambda n: (n.split()[-1].lower(), n.lower()))
    est = sum(1 for s in sections for m in s["m"] if m["est"])
    no_room = sum(1 for s in sections for m in s["m"] if m["s"] and not m["rm"])
    no_instr = sum(1 for s in sections if not s["i"])
    multi = sum(1 for s in sections if len(s["m"]) > 1)

    calendar = None
    if args.calendar:
        with open(args.calendar) as cf:
            calendar = json.load(cf)
    out = {
        "term": args.term,
        "termLabel": args.term_label or args.term,
        "generated": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "source": ", ".join(os.path.basename(p) for p in args.spreadsheet),
        "meetingSource": "scrape",
        "calendar": calendar,
        "faculty": faculty,
        "sections": sections,
    }
    outdir = os.path.dirname(os.path.abspath(args.out))
    if outdir:
        os.makedirs(outdir, exist_ok=True)
    with open(args.out, "w") as f:
        json.dump(out, f, separators=(",", ":"))

    print(f"{len(sections)} sections · {len(faculty)} instructors · {multi} multi-pattern -> {args.out}")
    print(f"  {est} meeting times with no end time" + ("   <-- these print with a ≈ flag" if est else ""))
    print(f"  {no_room} scheduled meetings with no room")
    print(f"  {no_instr} sections with no instructor (nobody's sheet will show them)")
    print(f"  {merged} instructor name variants merged")
    if calendar:
        print(f"  term {calendar.get('start')} to {calendar.get('end')}, "
              f"{len(calendar.get('closures') or [])} closure(s), {len(calendar.get('swaps') or [])} swap day(s)")
    for w in warnings[:15]:
        print("  ! " + w)
    if len(warnings) > 15:
        print(f"  ! …and {len(warnings) - 15} more")


if __name__ == "__main__":
    main()
